from django.shortcuts import render
from .quality import iffanalyse, iffimage
from .statistics import iffgraph
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from django.http import HttpResponse
from django.contrib import messages
from .models import bruikleenmodel
from .forms import bruikleenform

# Create your views here.
def home(request):
	return render(request, 'home.html')

def iff(request):
	return render(request, 'iff.html')

def ym(request):
	return render(request, 'ym.html')

def mm(request):
	return render(request, 'mm.html')

def iffstats(request):
	g001 = iffgraph.iff_g001()
	return render(request, 'iffstats.html', {'g001': g001})

def iffquality(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="#000.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'
	ws['A1'] = "list of sheet tab codes"
	ws.append(['sheet number', 'quality check'])
	ws.append(['#001', 'instellingsnaam != In Flanders Fields Museum'])
	ws.append(['#002', 'foutieve start objectnummer'])
	ws.append(['#003', 'foutieve lengte objectnummer'])
	ws.append(['#004', 'ontbrekende objectnaam'])
	ws.append(['#005', 'ontbrekende titel'])
	ws.append(['#006', 'foutieve titel'])
	ws.append(['#007', 'ontbrekende afmetingen'])
	ws.append(['#008', 'foutieve afmeting.eenheid'])
	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60
	iff_001 = iffanalyse.iff_001()
	df_001 = iff_001[0]
	iff_002 = iffanalyse.iff_002()
	df_002 = iff_002[0]
	iff_003 = iffanalyse.iff_003()
	df_003 = iff_003[0]
	iff_004 = iffanalyse.iff_004()
	df_004 = iff_004[0]
	iff_006 = iffanalyse.iff_006()
	df_005 = iff_006[0]
	df_006 = iff_006[2]
	iff_009 = iffanalyse.iff_009()
	df_007 = iff_009[0]
	df_008 = iff_009[1]
	if df_001.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001")
		rows = dataframe_to_rows(df_001, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002")
		rows = dataframe_to_rows(df_002, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_003.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#003")
		rows = dataframe_to_rows(df_003, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_004.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#004")
		rows = dataframe_to_rows(df_004, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_005.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#005")
		rows = dataframe_to_rows(df_005, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_006.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#006")
		rows = dataframe_to_rows(df_006, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_007.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#007")
		rows = dataframe_to_rows(df_007, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_008.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#008")
		rows = dataframe_to_rows(df_008, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response
			
def iffforms(request):
    if request.method == 'POST':
        form = bruikleenform(request.POST)
        if form.is_valid():
            form.save()
            form = bruikleenform()
        return render(request, 'iffforms.html', {'form': form})
    else:
        form = bruikleenform()
        return render(request, 'iffforms.html', {'form': form})
	
def iffimage(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="#000.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'
	ws['A1'] = "list of sheet tab codes"
	ws.append(['sheet number', 'quality check'])
	ws.append(['#001', 'records adlib te digitaliseren'])
	ws.append(['#002', 'objecten adlib te digitaliseren'])
	ws.append(['#003', 'fotos adlib te digitaliseren'])
	ws.append(['#004', 'documenten adlib te digitaliseren'])
	ws.append(['#005', 'records adlib, afbeelding te koppelen'])
	ws.append(['#006', 'beelden te registreren in adlib'])
	ws.append(['#007', 'HR beeld te maken van RAW'])
	ws.append(['#008', 'afgeleide te maken van HR/RAW beeld'])
	ws.append(['#009', 'geen HR/RAW beeld beschikbaar'])
	ws.append(['#010', 'HR < 300dpi'])
	ws.append(['#011', 'LR > 72dpi'])
	ws.append(['#012', '.tif beelden in LR'])
	ws.append(['#013', 'dubbele beelden'])
	df_001 = iffimage.iffi_001()
	df_002 = iffimage.iffi_002()
	df_003 = iffimage.iffi_003()
	df_004 = iffimage.iffi_004()
	df_005 = iffimage.iffi_005()
	df_006 = iffimage.iffi_006()
	df_007 = iffimage.iffi_007()
	df_008 = iffimage.iffi_008()
	df_009 = iffimage.iffi_009()
	df_010 = iffimage.iffi_010()
	df_011 = iffimage.iffi_011()
	df_012 = iffimage.iffi_012()
	df_013 = iffimage.iffi_013()

	if df_001.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001")
		rows = dataframe_to_rows(df_001, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002")
		rows = dataframe_to_rows(df_002, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_003.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#003")
		rows = dataframe_to_rows(df_003, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_004.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#004")
		rows = dataframe_to_rows(df_004, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_005.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#005")
		rows = dataframe_to_rows(df_005, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_006.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#006")
		rows = dataframe_to_rows(df_006, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_007.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#007")
		rows = dataframe_to_rows(df_007, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_008.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#008")
		rows = dataframe_to_rows(df_008, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_009.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#009")
		rows = dataframe_to_rows(df_009, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_010.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#010")
		rows = dataframe_to_rows(df_010, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_011.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#011")
		rows = dataframe_to_rows(df_011, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_012.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#012")
		rows = dataframe_to_rows(df_012, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_013.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#013")
		rows = dataframe_to_rows(df_013, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response