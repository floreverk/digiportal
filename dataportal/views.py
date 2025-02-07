from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from .quality import qualityiff, qualitymm, qualityym
from .statistics import statsiff, statsym, statsmm

# Create your views here.
def home(request):
    return render(request, 'home.html')

def iff(request):
	return render(request, 'iff.html')

def ym(request):
	return render(request, 'ym.html')

def mm(request):
	return render(request, 'mm.html')

def iffstats(request):
	g001 = statsiff.iff_g001()
	g002 = statsiff.iff_g002()
	g003 = statsiff.iff_g003()
	g004 = statsiff.iff_g004()
	g005 = statsiff.iff_g005()
	g006 = statsiff.iff_g006()
	g007 = statsiff.iff_g007()
	g008 = statsiff.iff_g008()
	g009 = statsiff.iff_g009()

	return render(request, 'iffstats.html', {'g001': g001, 'g002': g002, 'g003': g003, 'g004': g004, 
										  'g005': g005, 'g006': g006, 'g007': g007, 'g008': g008, 'g009': g009,})

def ymstats(request):
	g001 = statsym.ym_g001()
	g002 = statsym.ym_g002()
	g003 = statsym.ym_g003()
	g004 = statsym.ym_g004()
	g005 = statsym.ym_g005()
	g006 = statsym.ym_g006()
	g007 = statsym.ym_g007()
	g008 = statsym.ym_g008()
	g009 = statsym.ym_g009()

	return render(request, 'ymstats.html', {'g001': g001, 'g002': g002, 'g003': g003, 'g004': g004, 
										  'g005': g005, 'g006': g006, 'g007': g007, 'g008': g008, 'g009': g009,})

def mmstats(request):
	g001 = statsmm.mm_g001()
	g002 = statsmm.mm_g002()
	g003 = statsmm.mm_g003()
	g004 = statsmm.mm_g004()
	g006 = statsmm.mm_g006()
	g007 = statsmm.mm_g007()
	g008 = statsmm.mm_g008()

	return render(request, 'mmstats.html', {'g001': g001, 'g002': g002, 'g003': g003,'g004': g004, 
										  'g006': g006, 'g007': g007, 'g008': g008,})

def iffq001(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="identificatie.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
	    ['#001', 'INSTELLINGSNAAM'],
	    ['#001_01', 'instellingsnaam != In Flanders Fields Museum'],
	    ['#002', 'COLLECTIE'],
	    ['#002_01', 'collectie bevat lege occurences'],
	    ['#003', 'OBJECTNUMMER'],
	    ['#003_01', 'foutieve start objectnummer'],
	    ['#003_02', 'foutieve lengte objectnummer'],
	    ['#004', 'OBJECTCATEGORIE'],
	    ['#004_01', 'objectcategorie bevat lege occurences'],
	    ['#005', 'OBJECTNAAM'],
	    ['#005_01', 'objectnaam ontbreekt'],
	    ['#005_02', 'objectnaam start met hoofdletter'],
	    ['#005_03', 'objectnaam bevat lege occurences'],
	    ['#005_04', 'objectnaam komt twee maal voor'],
	    ['#006', 'TITEL'],
	    ['#006_01', 'titel ontbreekt'],
	    ['#006_02', 'foutieve start titel'],
	    ['#006_03', 'titel eindigt op punt/spatie'],
	    ['#006_04', 'titel is langer dan 250 karakters'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001', '#002', '#003', '#004', '#005', '#006']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	iff_q001 = qualityiff.iff_q001()
	df_001_01 = iff_q001[0]
	df_002_01 = iff_q001[1]
	df_003_01 = iff_q001[2]
	df_003_02 = iff_q001[3]
	df_004_01 = iff_q001[4]
	df_005_01 = iff_q001[5]
	df_005_02 = iff_q001[6]
	df_005_03 = iff_q001[7]
	df_005_04 = iff_q001[8]
	df_006_01 = iff_q001[9]
	df_006_02 = iff_q001[10]
	df_006_03 = iff_q001[11]
	df_006_04 = iff_q001[12]

	# Workbook sheets vullen
	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_01")
		rows = dataframe_to_rows(df_002_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_003_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#003_01")
		rows = dataframe_to_rows(df_003_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_003_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#003_02")
		rows = dataframe_to_rows(df_003_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_004_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#004_01")
		rows = dataframe_to_rows(df_004_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_005_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#005_01")
		rows = dataframe_to_rows(df_005_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_005_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#005_02")
		rows = dataframe_to_rows(df_005_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_005_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#005_03")
		rows = dataframe_to_rows(df_005_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_005_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#005_04")
		rows = dataframe_to_rows(df_005_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_006_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#006_01")
		rows = dataframe_to_rows(df_006_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_006_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#006_02")
		rows = dataframe_to_rows(df_006_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_006_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#006_03")
		rows = dataframe_to_rows(df_006_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_006_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#006_04")
		rows = dataframe_to_rows(df_006_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def iffq002(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="vervaardiging.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
	    ['#001', 'VERVAARDIGING'],
	    ['#001_01', 'vervaardiging periode is foutief'],
	    ['#001_02', 'vervaardiging datum begin precisie is foutief'],
	    ['#001_03', 'vervaardiging datum eind precisie is foutief'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	iff_q001 = qualityiff.iff_q002()
	df_001_01 = iff_q001[0]
	df_001_02 = iff_q001[1]
	df_001_03 = iff_q001[2]

	# Workbook sheets vullen
	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_03")
		rows = dataframe_to_rows(df_001_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def iffq003(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="fysiekekenmerken.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
	    ['#001', 'MATERIAAL'],
	    ['#001_01', 'lege occurences materiaal'],
	    ['#001_02', 'materiaal ontbreekt'],
	    ['#002', 'TECHNIEK'],
	    ['#002_01', 'lege occurences techniek'],
	    ['#002_02', 'techniek ontbreekt'],
		['#003', 'AFMETINGEN'],
	    ['#003_01', 'lege occurences afmetingen'],
	    ['#003_02', 'afmetingen ontbreken'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001', '#002', "#003"]:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	iff_q001 = qualityiff.iff_q003()
	df_001_01 = iff_q001[0]
	df_001_02 = iff_q001[1]
	df_002_01 = iff_q001[2]
	df_002_02 = iff_q001[3]
	df_003_01 = iff_q001[4]
	df_003_02 = iff_q001[5]

	# Workbook sheets vullen
	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_01")
		rows = dataframe_to_rows(df_002_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_02")
		rows = dataframe_to_rows(df_002_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_003_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#003_01")
		rows = dataframe_to_rows(df_003_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_003_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#003_02")
		rows = dataframe_to_rows(df_003_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def iffq004(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="iconoasso.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)

	data = [
	    ['#001', 'ICONOGRAFIE'],
	    ['#001_01', 'iconografie aanwezig maar soort ontbreekt'],
	    ['#001_02', 'lege occurences iconografie'],
	    ['#001_03', 'dubbele termen'],
	    ['#001_04', 'soort aanwezig maar iconografie ontbreekt'],
	    ['#001_05', 'foutieve iconografie soort'],
	    ['#002', 'ASSOCIATIES'],
	    ['#002_01', 'associatie aanwezig maar soort ontbreekt'],
	    ['#002_02', 'lege occurences associatie'],
	    ['#002_03', 'associatie periode is foutief'],
	    ['#002_04', 'dubbele termen'],
	    ['#002_05', 'soort aanwezig maar associatie ontbreekt'],
	    ['#002_06', 'foutieve associatie soort'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001', '#002']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	iff_q001 = qualityiff.iff_q004()
	df_001_01 = iff_q001[0]
	df_001_02 = iff_q001[1]
	df_001_03 = iff_q001[2]
	df_001_04 = iff_q001[3]
	df_001_05 = iff_q001[4]
	df_002_01 = iff_q001[5]
	df_002_02 = iff_q001[6]
	df_002_03 = iff_q001[7]
	df_002_04 = iff_q001[8]
	df_002_05 = iff_q001[9]
	df_002_06 = iff_q001[10]

	# Workbook sheets vullen
	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_03")
		rows = dataframe_to_rows(df_001_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_04")
		rows = dataframe_to_rows(df_001_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_05.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_05")
		rows = dataframe_to_rows(df_001_05, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_01")
		rows = dataframe_to_rows(df_002_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_02")
		rows = dataframe_to_rows(df_002_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_03")
		rows = dataframe_to_rows(df_002_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_04")
		rows = dataframe_to_rows(df_002_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_05.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_05")
		rows = dataframe_to_rows(df_002_05, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_06.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_06")
		rows = dataframe_to_rows(df_002_06, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def iffq005(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="rechten.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
	    ['#001', 'RECHTEN'],
	    ['#001_01', 'rechten type ontbreekt'],
	    ['#001_02', 'publiek domein zonder uitleg'],
	    ['#001_03', 'in copyright zonder einddatum'],
		['#001_04', 'rechten bijzonderheden foutief'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	iff_q001 = qualityiff.iff_q005()
	df_001_01 = iff_q001[0]
	df_001_02 = iff_q001[1]
	df_001_03 = iff_q001[2]
	df_001_04 = iff_q001[3]

	# Workbook sheets vullen
	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_03")
		rows = dataframe_to_rows(df_001_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_04")
		rows = dataframe_to_rows(df_001_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def iffq006(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="verwerving.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
	    ['#001', 'VERWERVING'],
	    ['#001_01', 'foutieve verwervingsmethode'],
	    ['#001_02', 'verwerving ontbreekt'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	iff_q001 = qualityiff.iff_q006()
	df_001_01 = iff_q001[0]
	df_001_02 = iff_q001[1]

	# Workbook sheets vullen
	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def ifft001(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="thesaurus.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
		['#001', 'TERM'],
		['#001_01', 'term soort = leeg'],
		['#001_02', 'term status =/ descriptor of non descriptor'],
		['#001_03', 'term start of eindigt met spaties'],
		['#002', 'BRON'],
		['#002_01', 'bron start of eindigt met spatie'],
		['#002_02', 'nummer start of eindigt met spatie'],
		['#002_03', 'status descriptor, bron en/of scopenote ontbreekt'],
		['#002_04', 'bron aanwezig, maar nummer ontbreekt'],
		['#002_05', 'nummer aanwezig, maar bron ontbreekt'],
		['#002_06', 'bron AAT, maar nummer =/ 9 digits'],
		['#002_07', 'bron Wikidata, maar nummer start niet met Q'],
		['#002_08', 'bron TGN, maar nummer =/ 7 digits'],
		['#002_09', 'foutieve bron'],
		['#002_10', 'non-descriptor termen komen voor bij records'],
		['#002_11', 'zelfde bron.nummers komen meermaals voor']
		]

	for row in data:
		ws.append(row)
		
	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001', '#002']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill
	
	t001 = qualityiff.t_001()
	df_001_01 = t001[0]
	df_001_02 = t001[1]
	df_001_03 = t001[2]
	t002 = qualityiff.t_002()
	df_002_01 = t002[0]
	df_002_02 = t002[1]
	df_002_03 = t002[2]
	df_002_04 = t002[3]
	df_002_05 = t002[4]
	df_002_06 = t002[5]
	df_002_07 = t002[6]
	df_002_08 = t002[7]
	df_002_09 = t002[8]
	df_002_10 = t002[9]
	df_002_11 = t002[10]

	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_03")
		rows = dataframe_to_rows(df_001_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_01")
		rows = dataframe_to_rows(df_002_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_02")
		rows = dataframe_to_rows(df_002_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_03")
		rows = dataframe_to_rows(df_002_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_04")
		rows = dataframe_to_rows(df_002_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_05.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_05")
		rows = dataframe_to_rows(df_002_05, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_06.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_06")
		rows = dataframe_to_rows(df_002_06, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_07.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_07")
		rows = dataframe_to_rows(df_002_07, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_08.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_08")
		rows = dataframe_to_rows(df_002_08, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_09.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_09")
		rows = dataframe_to_rows(df_002_09, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_10.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_10")
		rows = dataframe_to_rows(df_002_10, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_11.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_11")
		rows = dataframe_to_rows(df_002_11, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def iffb001(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="beeld.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
		['#001', 'UPLOADEN'],
	    ['#001_01', 'Afbeelding gevonden op A'],
		['#001_02', 'Record aan te maken in Adlib'],
		['#002', 'DIGITALISEREN'],
		['#002_01', 'Records te digitaliseren'],
		['#002_02', 'Objecten te digitaliseren'],
		['#002_03', 'Fotos te digitaliseren'],
		['#002_04', 'Documenten te digitaliseren'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001', '#002']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	iff_001 = qualityiff.iff_b001()
	df_001_01 = iff_001[0]
	df_001_02 = iff_001[1]
	df_002_01 = iff_001[2]
	df_002_02 = iff_001[3]
	df_002_03 = iff_001[4]
	df_002_04 = iff_001[5]

	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_01")
		rows = dataframe_to_rows(df_002_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_02")
		rows = dataframe_to_rows(df_002_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_03")
		rows = dataframe_to_rows(df_002_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_04")
		rows = dataframe_to_rows(df_002_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def iffb002(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="server.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
		['#001', 'RESOLUTIES'],
	    ['#001_01', 'RAW beelden zonder HR beeld'],
		['#001_02', 'HR/RAW beeld zonder LR beeld'],
		['#001_03', 'LR BEELD zonder HR/RAW beeld'],
		['#001_04', 'LR beeld in TIF'],
		['#002', 'BEELDEN'],
	    ['#002_01', 'dubbele bestanden'],
	    ['#002_02', 'foutieve mapnamen'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001', '#002']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	iff_001 = qualityiff.iff_b002()
	df_001_01 = iff_001[0]
	df_001_02 = iff_001[1]
	df_001_03 = iff_001[2]
	df_001_04 = iff_001[3]
	df_002_01 = iff_001[4]
	df_002_02 = iff_001[5]

	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_03")
		rows = dataframe_to_rows(df_001_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_04")
		rows = dataframe_to_rows(df_001_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_01")
		rows = dataframe_to_rows(df_002_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_02")
		rows = dataframe_to_rows(df_002_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def ymq001(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="identificatie.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
	    ['#001', 'INSTELLINGSNAAM'],
	    ['#001_01', 'instellingsnaam != Yper Museum, Onderwijs Museum, Stedelijk Museum of Museum Godshuis Belle'],
	    ['#002', 'COLLECTIE'],
	    ['#002_01', 'collectie bevat lege occurences'],
	    ['#003', 'OBJECTNUMMER'],
	    ['#003_01', 'foutieve start objectnummer'],
	    ['#004', 'OBJECTCATEGORIE'],
	    ['#004_01', 'objectcategorie bevat lege occurences'],
	    ['#005', 'OBJECTNAAM'],
	    ['#005_01', 'objectnaam ontbreekt'],
	    ['#005_02', 'objectnaam start met hoofdletter'],
	    ['#005_03', 'objectnaam bevat lege occurences'],
	    ['#005_04', 'objectnaam komt 2x voor'],
	    ['#006', 'TITEL'],
	    ['#006_01', 'titel ontbreekt'],
	    ['#006_02', 'foutieve start titel'],
	    ['#006_03', 'titel eindigt op punt/spatie'],
	    ['#006_04', 'titel is langer dan 250 karakters'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001', '#002', '#003', '#004', '#005', '#006']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	ym_q001 = qualityym.ym_q001()
	df_001_01 = ym_q001[0]
	df_002_01 = ym_q001[1]
	df_003_01 = ym_q001[2]
	df_004_01 = ym_q001[3]
	df_005_01 = ym_q001[4]
	df_005_02 = ym_q001[5]
	df_005_03 = ym_q001[6]
	df_005_04 = ym_q001[7]
	df_006_01 = ym_q001[8]
	df_006_02 = ym_q001[9]
	df_006_03 = ym_q001[10]
	df_006_04 = ym_q001[11]

	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_01")
		rows = dataframe_to_rows(df_002_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_003_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#003_01")
		rows = dataframe_to_rows(df_003_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_004_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#004_01")
		rows = dataframe_to_rows(df_004_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_005_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#005_01")
		rows = dataframe_to_rows(df_005_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_005_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#005_02")
		rows = dataframe_to_rows(df_005_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_005_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#005_03")
		rows = dataframe_to_rows(df_005_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_005_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#005_04")
		rows = dataframe_to_rows(df_005_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_006_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#006_01")
		rows = dataframe_to_rows(df_006_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_006_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#006_02")
		rows = dataframe_to_rows(df_006_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_006_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#006_03")
		rows = dataframe_to_rows(df_006_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_006_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#006_04")
		rows = dataframe_to_rows(df_006_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def ymq002(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="vervaardiging.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
	    ['#001', 'VERVAARDIGING'],
	    ['#001_01', 'vervaardiging periode is foutief'],
	    ['#001_02', 'vervaardiging datum begin precisie is foutief'],
	    ['#001_03', 'vervaardiging datum eind precisie is foutief'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	ym_q001 = qualityym.ym_q002()
	df_001_01 = ym_q001[0]
	df_001_02 = ym_q001[1]
	df_001_03 = ym_q001[2]

	# Workbook sheets vullen
	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_03")
		rows = dataframe_to_rows(df_001_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def ymq003(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="fysiekekenmerken.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
	    ['#001', 'MATERIAAL'],
	    ['#001_01', 'lege occurences materiaal'],
	    ['#001_02', 'materiaal ontbreekt'],
	    ['#002', 'TECHNIEK'],
	    ['#002_01', 'lege occurences techniek'],
	    ['#002_02', 'techniek ontbreekt'],
		['#003', 'AFMETINGEN'],
	    ['#003_01', 'lege occurences afmetingen'],
	    ['#003_02', 'afmetingen ontbreken'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001', '#002', "#003"]:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	ym_q001 = qualityym.ym_q003()
	df_001_01 = ym_q001[0]
	df_001_02 = ym_q001[1]
	df_002_01 = ym_q001[2]
	df_002_02 = ym_q001[3]
	df_003_01 = ym_q001[4]
	df_003_02 = ym_q001[5]

	# Workbook sheets vullen
	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_01")
		rows = dataframe_to_rows(df_002_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_02")
		rows = dataframe_to_rows(df_002_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_003_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#003_01")
		rows = dataframe_to_rows(df_003_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_003_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#003_02")
		rows = dataframe_to_rows(df_003_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def ymq004(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="iconoasso.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
	    ['#001', 'ICONOGRAFIE'],
	    ['#001_01', 'iconografie aanwezig maar soort ontbreekt'],
	    ['#001_02', 'lege occurences iconografie'],
	    ['#001_03', 'dubbele termen'],
	    ['#001_04', 'soort aanwezig maar iconografie ontbreekt'],
	    ['#001_05', 'foutieve iconografie soort'],
	    ['#002', 'ASSOCIATIES'],
	    ['#002_01', 'associatie aanwezig maar soort ontbreekt'],
	    ['#002_02', 'lege occurences associatie'],
	    ['#002_03', 'associatie periode is foutief'],
	    ['#002_04', 'dubbele termen'],
	    ['#002_05', 'soort aanwezig maar associatie ontbreekt'],
	    ['#002_06', 'foutieve associatie soort'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001', '#002']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	ym_q001 = qualityym.ym_q004()
	df_001_01 = ym_q001[0]
	df_001_02 = ym_q001[1]
	df_001_03 = ym_q001[2]
	df_001_04 = ym_q001[3]
	df_001_05 = ym_q001[4]
	df_002_01 = ym_q001[5]
	df_002_02 = ym_q001[6]
	df_002_03 = ym_q001[7]
	df_002_04 = ym_q001[8]
	df_002_05 = ym_q001[9]
	df_002_06 = ym_q001[10]

	# Workbook sheets vullen
	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_03")
		rows = dataframe_to_rows(df_001_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_04")
		rows = dataframe_to_rows(df_001_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_05.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_05")
		rows = dataframe_to_rows(df_001_05, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_01")
		rows = dataframe_to_rows(df_002_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_02")
		rows = dataframe_to_rows(df_002_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_03")
		rows = dataframe_to_rows(df_002_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_04")
		rows = dataframe_to_rows(df_002_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_05.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_05")
		rows = dataframe_to_rows(df_002_05, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_06.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_06")
		rows = dataframe_to_rows(df_002_06, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def ymq005(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="rechten.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
	    ['#001', 'RECHTEN'],
	    ['#001_01', 'rechten type ontbreekt'],
	    ['#001_02', 'publiek domein zonder uitleg'],
	    ['#001_03', 'in copyright zonder einddatum'],
		['#001_04', 'rechten bijzonderheden foutief'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	ym_q001 = qualityym.ym_q005()
	df_001_01 = ym_q001[0]
	df_001_02 = ym_q001[1]
	df_001_03 = ym_q001[2]
	df_001_04 = ym_q001[3]

	# Workbook sheets vullen
	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_03")
		rows = dataframe_to_rows(df_001_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_04")
		rows = dataframe_to_rows(df_001_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def ymq006(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="verwerving.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
	    ['#001', 'VERWERVING'],
	    ['#001_01', 'foutieve verwervingsmethode'],
	    ['#001_02', 'verwerving ontbreekt'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	ym_q001 = qualityym.ym_q006()
	df_001_01 = ym_q001[0]
	df_001_02 = ym_q001[1]

	# Workbook sheets vullen
	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def ymt001(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="thesaurus.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
		['#001', 'TERM'],
		['#001_01', 'term soort = leeg'],
		['#001_02', 'term status =/ descriptor of non descriptor'],
		['#001_03', 'term start of eindigt met spaties'],
		['#002', 'BRON'],
		['#002_01', 'bron start of eindigt met spatie'],
		['#002_02', 'nummer start of eindigt met spatie'],
		['#002_03', 'status descriptor, bron en/of scopenote ontbreekt'],
		['#002_04', 'bron aanwezig, maar nummer ontbreekt'],
		['#002_05', 'nummer aanwezig, maar bron ontbreekt'],
		['#002_06', 'bron AAT, maar nummer =/ 9 digits'],
		['#002_07', 'bron Wikidata, maar nummer start niet met Q'],
		['#002_08', 'bron TGN, maar nummer =/ 7 digits'],
		['#002_09', 'foutieve bron'],
		['#002_10', 'non-descriptor termen komen voor bij records']
		]

	for row in data:
		ws.append(row)
		
	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001', '#002']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill
	
	t001 = qualityym.ym_t001()
	df_001_01 = t001[0]
	df_001_02 = t001[1]
	df_001_03 = t001[2]
	t002 = qualityym.ym_t002()
	df_002_01 = t002[0]
	df_002_02 = t002[1]
	df_002_03 = t002[2]
	df_002_04 = t002[3]
	df_002_05 = t002[4]
	df_002_06 = t002[5]
	df_002_07 = t002[6]
	df_002_08 = t002[7]
	df_002_09 = t002[8]
	df_002_10 = t002[9]

	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_03")
		rows = dataframe_to_rows(df_001_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_01")
		rows = dataframe_to_rows(df_002_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_02")
		rows = dataframe_to_rows(df_002_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_03")
		rows = dataframe_to_rows(df_002_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_04")
		rows = dataframe_to_rows(df_002_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_05.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_05")
		rows = dataframe_to_rows(df_002_05, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_06.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_06")
		rows = dataframe_to_rows(df_002_06, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_07.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_07")
		rows = dataframe_to_rows(df_002_07, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_08.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_08")
		rows = dataframe_to_rows(df_002_08, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_09.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_09")
		rows = dataframe_to_rows(df_002_09, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_10.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_10")
		rows = dataframe_to_rows(df_002_10, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def ymb001(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="beeld.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
		['#001', 'UPLOADEN'],
	    ['#001_01', 'Afbeelding gevonden op A'],
		['#001_02', 'Record aan te maken in Adlib'],
		['#002', 'DIGITALISEREN'],
		['#002_01', 'Records te digitaliseren'],
		['#002_02', 'Objecten te digitaliseren'],
		['#002_03', 'Fotos te digitaliseren'],
		['#002_04', 'Documenten te digitaliseren'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001', '#002']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	ym_001 = qualityym.ym_b001()
	df_001_01 = ym_001[0]
	df_001_02 = ym_001[1]
	df_002_01 = ym_001[2]
	df_002_02 = ym_001[3]
	df_002_03 = ym_001[4]
	df_002_04 = ym_001[5]

	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_01")
		rows = dataframe_to_rows(df_002_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_02")
		rows = dataframe_to_rows(df_002_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_03")
		rows = dataframe_to_rows(df_002_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_04")
		rows = dataframe_to_rows(df_002_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def ymb002(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="server.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
		['#001', 'RESOLUTIES'],
	    ['#001_01', 'RAW beelden zonder HR beeld'],
		['#001_02', 'HR/RAW beeld zonder LR beeld'],
		['#001_03', 'LR BEELD zonder HR/RAW beeld'],
		['#001_04', 'LR beeld in TIF'],
		['#002', 'BEELDEN'],
	    ['#002_01', 'dubbele bestanden'],
	    ['#002_02', 'foutieve mapnamen'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001', '#002']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	ym_001 = qualityym.ym_b002()
	df_001_01 = ym_001[0]
	df_001_02 = ym_001[1]
	df_001_03 = ym_001[2]
	df_001_04 = ym_001[3]
	df_002_01 = ym_001[4]
	df_002_02 = ym_001[5]

	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_03")
		rows = dataframe_to_rows(df_001_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_04")
		rows = dataframe_to_rows(df_001_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_01")
		rows = dataframe_to_rows(df_002_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_02")
		rows = dataframe_to_rows(df_002_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def mmq001(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="identificatie.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
	    ['#001', 'INSTELLINGSNAAM'],
	    ['#001_01', 'instellingsnaam != In Flanders Fields Museum'],
	    ['#002', 'COLLECTIE'],
	    ['#002_01', 'collectie bevat lege occurences'],
	    ['#003', 'OBJECTNUMMER'],
	    ['#003_01', 'foutieve start objectnummer'],
	    ['#003_02', 'foutieve lengte objectnummer'],
	    ['#004', 'OBJECTCATEGORIE'],
	    ['#004_01', 'objectcategorie bevat lege occurences'],
	    ['#005', 'OBJECTNAAM'],
	    ['#005_01', 'objectnaam ontbreekt'],
	    ['#005_02', 'objectnaam start met hoofdletter'],
	    ['#005_03', 'objectnaam bevat lege occurences'],
	    ['#005_04', 'objectnaam komt 2x voor'],
	    ['#006', 'TITEL'],
	    ['#006_01', 'titel ontbreekt'],
	    ['#006_02', 'foutieve start titel'],
	    ['#006_03', 'titel eindigt op punt/spatie'],
	    ['#006_04', 'titel is langer dan 250 karakters'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001', '#002', '#003', '#004', '#005', '#006']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	mm_q001 = qualitymm.mm_q001()
	df_001_01 = mm_q001[0]
	df_002_01 = mm_q001[1]
	df_003_01 = mm_q001[2]
	df_003_02 = mm_q001[3]
	df_004_01 = mm_q001[4]
	df_005_01 = mm_q001[5]
	df_005_02 = mm_q001[6]
	df_005_03 = mm_q001[7]
	df_005_04 = mm_q001[8]
	df_006_01 = mm_q001[9]
	df_006_02 = mm_q001[10]
	df_006_03 = mm_q001[11]
	df_006_04 = mm_q001[12]

	# Workbook sheets vullen
	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_01")
		rows = dataframe_to_rows(df_002_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_003_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#003_01")
		rows = dataframe_to_rows(df_003_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_003_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#003_02")
		rows = dataframe_to_rows(df_003_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_004_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#004_01")
		rows = dataframe_to_rows(df_004_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_005_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#005_01")
		rows = dataframe_to_rows(df_005_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_005_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#005_02")
		rows = dataframe_to_rows(df_005_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_005_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#005_03")
		rows = dataframe_to_rows(df_005_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_005_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#005_04")
		rows = dataframe_to_rows(df_005_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_006_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#006_01")
		rows = dataframe_to_rows(df_006_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_006_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#006_02")
		rows = dataframe_to_rows(df_006_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_006_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#006_03")
		rows = dataframe_to_rows(df_006_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_006_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#006_04")
		rows = dataframe_to_rows(df_006_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def mmq002(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="vervaardiging.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
	    ['#001', 'VERVAARDIGING'],
	    ['#001_01', 'vervaardiging periode is foutief'],
	    ['#001_02', 'vervaardiging datum begin precisie is foutief'],
	    ['#001_03', 'vervaardiging datum eind precisie is foutief'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	mm_q001 = qualitymm.mm_q002()
	df_001_01 = mm_q001[0]
	df_001_02 = mm_q001[1]
	df_001_03 = mm_q001[2]

	# Workbook sheets vullen
	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_03")
		rows = dataframe_to_rows(df_001_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def mmq003(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="fysiekekenmerken.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
	    ['#001', 'MATERIAAL'],
	    ['#001_01', 'lege occurences materiaal'],
	    ['#001_02', 'materiaal ontbreekt'],
	    ['#002', 'TECHNIEK'],
	    ['#002_01', 'lege occurences techniek'],
	    ['#002_02', 'techniek ontbreekt'],
		['#003', 'AFMETINGEN'],
	    ['#003_01', 'lege occurences afmetingen'],
	    ['#003_02', 'afmetingen ontbreken'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001', '#002', "#003"]:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	mm_q001 = qualitymm.mm_q003()
	df_001_01 = mm_q001[0]
	df_001_02 = mm_q001[1]
	df_002_01 = mm_q001[2]
	df_002_02 = mm_q001[3]
	df_003_01 = mm_q001[4]
	df_003_02 = mm_q001[5]

	# Workbook sheets vullen
	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_01")
		rows = dataframe_to_rows(df_002_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_02")
		rows = dataframe_to_rows(df_002_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_003_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#003_01")
		rows = dataframe_to_rows(df_003_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_003_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#003_02")
		rows = dataframe_to_rows(df_003_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def mmq004(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="iconoasso.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
	    ['#001', 'ICONOGRAFIE'],
	    ['#001_01', 'iconografie aanwezig maar soort ontbreekt'],
	    ['#001_02', 'lege occurences iconografie'],
	    ['#001_03', 'dubbele termen'],
	    ['#001_04', 'soort aanwezig maar iconografie ontbreekt'],
	    ['#001_05', 'foutieve soort iconografie'],
	    ['#002', 'ASSOCIATIES'],
	    ['#002_01', 'associatie aanwezig maar soort ontbreekt'],
	    ['#002_02', 'lege occurences associatie'],
	    ['#002_03', 'associatie periode is foutief'],
	    ['#002_04', 'dubbele termen'],
	    ['#002_05', 'soort aanwezig maar associatie ontbreekt'],
	    ['#002_06', 'foutieve soort associatie'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001', '#002']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	mm_q001 = qualitymm.mm_q004()
	df_001_01 = mm_q001[0]
	df_001_02 = mm_q001[1]
	df_001_03 = mm_q001[2]
	df_001_04 = mm_q001[3]
	df_001_05 = mm_q001[4]
	df_002_01 = mm_q001[5]
	df_002_02 = mm_q001[6]
	df_002_03 = mm_q001[7]
	df_002_04 = mm_q001[8]
	df_002_05 = mm_q001[9]
	df_002_06 = mm_q001[10]

	# Workbook sheets vullen
	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_03")
		rows = dataframe_to_rows(df_001_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_04")
		rows = dataframe_to_rows(df_001_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_05.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_05")
		rows = dataframe_to_rows(df_001_05, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_01")
		rows = dataframe_to_rows(df_002_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_02")
		rows = dataframe_to_rows(df_002_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_03")
		rows = dataframe_to_rows(df_002_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_04")
		rows = dataframe_to_rows(df_002_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_05.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_05")
		rows = dataframe_to_rows(df_002_05, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_06.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_06")
		rows = dataframe_to_rows(df_002_06, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def mmq005(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="rechten.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
	    ['#001', 'RECHTEN'],
	    ['#001_01', 'rechten type ontbreekt'],
	    ['#001_02', 'publiek domein zonder uitleg'],
	    ['#001_03', 'in copyright zonder einddatum'],
		['#001_04', 'rechten bijzonderheden foutief'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	mm_q001 = qualitymm.mm_q005()
	df_001_01 = mm_q001[0]
	df_001_02 = mm_q001[1]
	df_001_03 = mm_q001[2]
	df_001_04 = mm_q001[3]

	# Workbook sheets vullen
	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_03")
		rows = dataframe_to_rows(df_001_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_04")
		rows = dataframe_to_rows(df_001_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def mmq006(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="verwerving.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
	    ['#001', 'VERWERVING'],
	    ['#001_01', 'foutieve verwervingsmethode'],
	    ['#001_02', 'verwerving ontbreekt'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	mm_q001 = qualitymm.mm_q006()
	df_001_01 = mm_q001[0]
	df_001_02 = mm_q001[1]

	# Workbook sheets vullen
	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def mmt001(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="thesaurus.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
		['#001', 'TERM'],
		['#001_01', 'term soort = leeg'],
		['#001_02', 'term status =/ descriptor of non descriptor'],
		['#001_03', 'term start of eindigt met spaties'],
		['#002', 'BRON'],
		['#002_01', 'bron start of eindigt met spatie'],
		['#002_02', 'nummer start of eindigt met spatie'],
		['#002_03', 'status descriptor, bron en/of scopenote ontbreekt'],
		['#002_04', 'bron aanwezig, maar nummer ontbreekt'],
		['#002_05', 'nummer aanwezig, maar bron ontbreekt'],
		['#002_06', 'bron AAT, maar nummer =/ 9 digits'],
		['#002_07', 'bron Wikidata, maar nummer start niet met Q'],
		['#002_08', 'bron TGN, maar nummer =/ 7 digits'],
		['#002_09', 'foutieve bron'],
		['#002_10', 'non-descriptor termen komen voor bij records']
		]

	for row in data:
		ws.append(row)
		
	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001', '#002']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill
	
	t001 = qualitymm.mm_t001()
	df_001_01 = t001[0]
	df_001_02 = t001[1]
	df_001_03 = t001[2]
	t002 = qualitymm.mm_t002()
	df_002_01 = t002[0]
	df_002_02 = t002[1]
	df_002_03 = t002[2]
	df_002_04 = t002[3]
	df_002_05 = t002[4]
	df_002_06 = t002[5]
	df_002_07 = t002[6]
	df_002_08 = t002[7]
	df_002_09 = t002[8]
	df_002_10 = t002[9]

	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_03")
		rows = dataframe_to_rows(df_001_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_01")
		rows = dataframe_to_rows(df_002_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_02")
		rows = dataframe_to_rows(df_002_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_03")
		rows = dataframe_to_rows(df_002_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_04")
		rows = dataframe_to_rows(df_002_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_05.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_05")
		rows = dataframe_to_rows(df_002_05, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_06.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_06")
		rows = dataframe_to_rows(df_002_06, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_07.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_07")
		rows = dataframe_to_rows(df_002_07, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_08.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_08")
		rows = dataframe_to_rows(df_002_08, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_09.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_09")
		rows = dataframe_to_rows(df_002_09, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_10.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_10")
		rows = dataframe_to_rows(df_002_10, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def mmb001(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="beeld.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
		['#001', 'UPLOADEN'],
	    ['#001_01', 'Afbeelding gevonden op A'],
		['#001_02', 'Record aan te maken in Adlib'],
		['#002', 'DIGITALISEREN'],
		['#002_01', 'Records te digitaliseren'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001', '#002']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	mm_001 = qualitymm.mm_b001()
	df_001_01 = mm_001[0]
	df_001_02 = mm_001[1]
	df_002_01 = mm_001[2]

	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_01")
		rows = dataframe_to_rows(df_002_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response

def mmb002(request):
	response = HttpResponse(content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="server.xlsx"'
	wb = Workbook()
	ws = wb.active
	ws.title = 'Info'

	ws['A1'] = "LIST OF SHEET CODES"
	ws.merge_cells('A1:B1')
	header_font = Font(color="FFFFFF", bold=True, size=16)
	header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'].font = header_font
	ws['A1'].fill = header_fill
	ws['A1'].alignment = header_alignment

	ws.append(['SHEETNUMBER', 'QUALITYCHECK'])
	for cell in ws["2:2"]:
		cell.font = Font(bold=True, size=16)
	
	data = [
		['#001', 'RESOLUTIES'],
	    ['#001_01', 'RAW beelden zonder HR beeld'],
		['#001_02', 'HR/RAW beeld zonder LR beeld'],
		['#001_03', 'LR BEELD zonder HR/RAW beeld'],
		['#001_04', 'LR beeld in TIF'],
		['#002', 'BEELDEN'],
	    ['#002_01', 'dubbele bestanden'],
	    ['#002_02', 'foutieve mapnamen'],
	    ]

	for row in data:
		ws.append(row)

	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 60

	highlight_font = Font(color="FFFFFF", bold=True) # Witte tekst en vet
	highlight_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Zwarte achtergrond

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
		if row[0].value in ['#001', '#002']:  # Check voor specifieke termen
			for cell in row:
				cell.font = highlight_font
				cell.fill = highlight_fill

    # Data selecteren
	mm_001 = qualitymm.mm_b002()
	df_001_01 = mm_001[0]
	df_001_02 = mm_001[1]
	df_001_03 = mm_001[2]
	df_001_04 = mm_001[3]
	df_002_01 = mm_001[4]
	df_002_02 = mm_001[5]

	if df_001_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_01")
		rows = dataframe_to_rows(df_001_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_02")
		rows = dataframe_to_rows(df_001_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_03.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_03")
		rows = dataframe_to_rows(df_001_03, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_001_04.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#001_04")
		rows = dataframe_to_rows(df_001_04, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_01.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_01")
		rows = dataframe_to_rows(df_002_01, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	if df_002_02.empty == True:
		print('empty dataframe')
	else:
		ws = wb.create_sheet("#002_02")
		rows = dataframe_to_rows(df_002_02, index=False)
		for r_idx, row in enumerate(rows, 1):
			for c_idx, value in enumerate(row, 1):
				ws.cell(row=r_idx, column=c_idx, value=value)
	wb.save(response)
	return response
